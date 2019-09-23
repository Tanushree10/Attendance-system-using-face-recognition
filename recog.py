import sqlite3
import face_recognition
import cv2
import dlib
import numpy as np
import os
from openpyxl import Workbook, load_workbook
import time
from train import *
cam = cv2.VideoCapture(0)

#get current date
currentDate = time.strftime("%d_%m_%y")
  
# workbook object is created 
wb_obj = load_workbook(currentDate+".xlsx") 
sheet_obj = wb_obj.active 
m_row = sheet_obj.max_row 

def getProfile(id):
   
    conn=sqlite3.connect('face_database.db')
    
    cmd='select * from attendance where ID='+str(id)
    cursor=conn.execute(cmd)
    
    profile=None
    for row in cursor:
       profile=row
    conn.close()
    return profile  


ids = os.listdir('images')
classes = []

for i in ids:
    encodings = []
    for j in os.listdir('images/'+i):
        img = face_recognition.load_image_file("images/"+i+"/"+str(j))
        img_encoding = face_recognition.face_encodings(img)[0]
        
        encodings.append(img_encoding)
    classes.append(encodings)

detector = dlib.get_frontal_face_detector()
color_green = (0,255,0)
line_width = 3

print('open')
while True:
    ret_val, img = cam.read()
    #rgb_image = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    tracker = cv2.TrackerBoosting_create()
    dets = detector(img)
    
    for det in dets:
        
        crop_img = img[det.left():det.right(), det.top():det.bottom()]
        
    
        bbox = (det.left()-5, det.top()-5, det.right() - det.left()+10 ,det.bottom() - det.top()+10)

        unknown_picture = face_recognition.face_locations(img)
        
        flag = 0
        try:
            unknown_face_encoding = face_recognition.face_encodings(img,unknown_picture)[0]
            #print(unknown_face_encoding)
            for i in range (5):
                cv2.rectangle(img,(det.left(), det.top()), (det.right(), det.bottom()), color_green, line_width)
                for j in range (len(ids)):
                    
                    results1 = face_recognition.compare_faces([classes[j][i]], unknown_face_encoding, tolerance = 0.50)
                    if results1[0] == True:
                        #print(results1)
                        break

                       
                if results1[0] == True:
                    flag = 1
                    profile=getProfile(ids[j])
                    print(profile[1])
                   # if(profile!=None):
                
        
                    cv2.putText(img,profile[1], (det.left(), det.top()), cv2.FONT_HERSHEY_SIMPLEX, 0.75,(0,0,255),2)
                    for k in range(1, m_row + 1): 
                        cell_obj = sheet_obj.cell(row = k, column = 1)
                        if (cell_obj.value==profile[0]):
                            b=k
                            wb = load_workbook(filename = currentDate+".xlsx")
                            ws1 = wb.active
                            c=ws1.cell(row=b ,column=4)
                            c.value="Present"
                            wb.save(currentDate+".xlsx")
                    break
            if not flag:
                cv2.putText(img, "none", (det.left(), det.top()), cv2.FONT_HERSHEY_SIMPLEX, 0.75,(0,0,255),2)
        except (RuntimeError, TypeError, NameError, IndexError):
            print("no image found")

        
    cv2.imshow('LiveStream', img)
        
   
    # Exit if ESC pressed    
    k = cv2.waitKey(1) & 0xff
    if k == 27 : break
