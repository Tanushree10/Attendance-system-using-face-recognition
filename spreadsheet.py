from openpyxl import Workbook, load_workbook
import time
import os
import sqlite3

#database connection
conn = sqlite3.connect('face_database.db')
c = conn.cursor()

#get current date
currentDate = time.strftime("%d_%m_%y")

#create a workbook and add a worksheet
if(os.path.exists(currentDate+".xlsx")):
    wb = load_workbook(filename = currentDate+".xlsx")
else:
    wb = Workbook()
    dest_filename = currentDate+'.xlsx'
    c.execute("SELECT * FROM attendance ORDER BY ID ASC")
    
    #creating worksheet and giving names to column
    ws1 = wb.active
    ws1.title = 'IIIT'
    ws1.append(('ID', 'USERNAME','BRANCH', currentDate))
    ws1.append(('', '', '',''))

    #entering students information from database
    while True:
        a = c.fetchone()
        if a == None:
            break
        else:
            print(a)
            ws1.append((a[0], a[1],a[2]))

    #saving the file
    wb.save(filename = dest_filename)
